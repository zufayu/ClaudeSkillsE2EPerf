#!/usr/bin/env python3
"""
Wrapper around ATOM's parse_trace.py with --target-bs and multi-EP support.

ATOM's parse_trace.py assumes all transformer layer modules are direct children
of a single CompiledFxGraph block inside capture_graph. With EP>1, torch.compile
splits the model into multiple CompiledFxGraph blocks (one per few layers), and
parse_trace only sees modules from the first block → incomplete breakdown.

This wrapper fixes three issues:
  1. Target-BS: selects steady-state decode events instead of the first one
  2. Multi-EP: flattens the capture_graph hierarchy so all modules from all
     CompiledFxGraph blocks appear as direct children of capture_graph
  3. Wide-sample aggregation (R8d): aggregates over N steady-state decodes
     (default --skip-warmup 5 --max-steps 20 → ~600 samples/op) instead of
     picking a single decode. Output xlsx adds median_us / p95_us / n_steps
     columns. Methodology mirrors B300/B200 trace_layer_detail.py for
     apples-to-apples cross-platform comparison. Pass --max-steps 1 with
     --skip-ratio for back-compat single-decode behavior.

Usage:
    python3 scripts/run_parse_trace.py <trace.json.gz> [--layer N] [--target-bs N]

    # R8d default: skip 5 warmup decodes, aggregate next 20 (~600 samples/op):
    python3 scripts/run_parse_trace.py trace.json.gz --layer 40 --target-bs 64

    # Back-compat single-decode (median position):
    python3 scripts/run_parse_trace.py trace.json.gz --target-bs 64 \\
        --max-steps 1 --skip-ratio 0.5

Output:
    decode_breakdown_c64.xlsx   (kernel breakdown at target bs; columns:
                                 cpu_module, gpu_kernel, avg_us, median_us,
                                 p95_us, std_us, n_steps, pct%)
    decode_breakdown_c64.csv    (sister CSV in same schema for diff scripts)
    prefill_breakdown_c64.xlsx  (prefill breakdown, unchanged from upstream)
    decode_per_step_c64/        (per-step intermediate xlsx, when N > 1)

Requires ATOM's tools/ on sys.path. Set ATOM_TOOLS env var if not at default.
"""

import argparse
import copy
import re
import sys
import os

# Add ATOM tools to path
_SEARCH_PATHS = [
    os.environ.get("ATOM_TOOLS", ""),
    "/app/ATOM/tools",
    os.path.expanduser("~/ATOM/tools"),
    "/home/kqian/ATOM/tools",
]
for p in _SEARCH_PATHS:
    if p and os.path.isfile(os.path.join(p, "parse_trace.py")):
        sys.path.insert(0, p)
        break

try:
    import parse_trace
except ImportError:
    print("ERROR: Cannot import parse_trace. Set ATOM_TOOLS=/path/to/ATOM/tools")
    sys.exit(1)


def select_decodes_bs(events, target_bs=None, skip_warmup=5, max_steps=20,
                       skip_ratio=None):
    """Find target bs and a list of steady-state decode events at that bs.

    R8d wide-sample default: skip first `skip_warmup` decodes (cold caches /
    JIT compile), then take next `max_steps` for ~600 (decode×layer) samples
    per kernel — matches B300/B200 trace_layer_detail.py convention so cross-
    platform comparison is apples-to-apples. Set max_steps=1 with skip_ratio
    to reproduce the old single-decode picker behavior for back-compat.

    Returns (list_of_decodes, target_bs).
    """
    decodes = sorted(
        [
            e for e in events
            if e.get("name", "").startswith("decode[")
            and e.get("ph") == "X"
            and e.get("cat") == "gpu_user_annotation"
        ],
        key=lambda x: x["ts"],
    )
    if not decodes:
        return [], None

    bs_counts = {}
    decodes_by_bs = {}
    for d in decodes:
        m = re.search(r"bs=(\d+)", d.get("name", ""))
        if m:
            bs = int(m.group(1))
            bs_counts[bs] = bs_counts.get(bs, 0) + 1
            decodes_by_bs.setdefault(bs, []).append(d)

    print(f"Decode bs distribution: {dict(sorted(bs_counts.items()))}")

    if target_bs is None:
        target_bs = max(bs_counts, key=lambda b: bs_counts[b])
        print(f"Auto-selected target_bs={target_bs} ({bs_counts[target_bs]} events)")
    else:
        print(f"--target-bs={target_bs} ({bs_counts.get(target_bs, 0)} events)")

    candidates = decodes_by_bs.get(target_bs, [])
    if not candidates:
        return [], target_bs

    # Back-compat single-decode mode: explicit skip_ratio + max_steps=1
    if skip_ratio is not None and max_steps == 1:
        idx = min(int(len(candidates) * skip_ratio), len(candidates) - 1)
        selected = [candidates[idx]]
        dur_ms = selected[0].get("dur", 0) / 1000
        print(f"  [single-decode mode] picked #{idx}/{len(candidates)} "
              f"(skip_ratio={skip_ratio}, dur={dur_ms:.2f}ms)")
        return selected, target_bs

    # Default R8d wide-sample mode
    if len(candidates) <= skip_warmup:
        print(f"  WARN: only {len(candidates)} decodes, can't skip {skip_warmup} warmup; using all")
        selected = candidates
    else:
        end = skip_warmup + max_steps if max_steps > 0 else len(candidates)
        selected = candidates[skip_warmup:end]
    avg_dur_ms = sum(d.get("dur", 0) for d in selected) / len(selected) / 1000
    print(f"  Selected {len(selected)} steady-state decodes "
          f"(skipped first {skip_warmup} warmup, max {max_steps}); "
          f"avg dur={avg_dur_ms:.2f}ms")
    return selected, target_bs


# Back-compat shim: old single-decode picker (kept for any external callers).
def select_decode_bs(events, target_bs=None, skip_ratio=0.5):
    """DEPRECATED: returns one decode. Use select_decodes_bs going forward."""
    decodes, bs = select_decodes_bs(events, target_bs,
                                     skip_warmup=0, max_steps=1,
                                     skip_ratio=skip_ratio)
    return (decodes[0] if decodes else None), bs


def merge_decode_xlsx(per_step_paths, output_xlsx):
    """Aggregate N per-step decode xlsx files into a single wide-sample xlsx.

    Reads each step's "decode" sheet, groups by (cpu_module, gpu_kernel),
    and emits mean / median / p95 / std / N per kernel — same schema convention
    as B300 trace_layer_detail.py output (apples-to-apples cross-platform).

    Output sheet "decode":
      cpu_module | gpu_kernel | avg_us | median_us | p95_us | std_us | n_steps | pct%
    """
    from openpyxl import load_workbook, Workbook
    from collections import defaultdict
    from statistics import pstdev

    if not per_step_paths:
        print("merge_decode_xlsx: no input files"); return

    pooled = defaultdict(list)   # (cpu_module, gpu_kernel) -> [duration_us, ...]
    order = []                   # preserve first-occurrence order
    seen = set()
    for p in per_step_paths:
        if not os.path.exists(p):
            print(f"  WARN: missing {p}"); continue
        try:
            wb = load_workbook(p, read_only=True, data_only=True)
        except Exception as e:
            print(f"  WARN: failed to load {p}: {e}"); continue
        ws = wb["decode"] if "decode" in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            wb.close(); continue
        # Skip header (row 0); skip TOTAL row
        for r in rows[1:]:
            if not r or r[0] in (None, "TOTAL"):
                continue
            cpu_mod, gpu_kernel, dur = r[0], r[1], r[2]
            if dur is None or gpu_kernel is None:
                continue
            try:
                dur = float(dur)
            except (TypeError, ValueError):
                continue
            key = (str(cpu_mod or ""), str(gpu_kernel))
            pooled[key].append(dur)
            if key not in seen:
                seen.add(key); order.append(key)
        wb.close()

    if not pooled:
        print("merge_decode_xlsx: no kernel data extracted"); return

    def percentile(sorted_vals, p):
        if not sorted_vals: return 0.0
        k = (len(sorted_vals) - 1) * p / 100
        lo = int(k); hi = min(lo + 1, len(sorted_vals) - 1)
        return sorted_vals[lo] + (sorted_vals[hi] - sorted_vals[lo]) * (k - lo)

    means = {k: sum(v)/len(v) for k, v in pooled.items()}
    total_avg = sum(means.values())

    out_wb = Workbook(); ws = out_wb.active; ws.title = "decode"
    ws.append(["cpu_module", "gpu_kernel", "avg_us", "median_us", "p95_us",
               "std_us", "n_steps", "pct%"])
    for key in order:
        durs = pooled[key]
        durs_sorted = sorted(durs)
        avg = means[key]
        med = percentile(durs_sorted, 50)
        p95 = percentile(durs_sorted, 95)
        std = pstdev(durs) if len(durs) > 1 else 0.0
        pct = 100 * avg / total_avg if total_avg > 0 else 0
        ws.append([key[0], key[1],
                   round(avg, 3), round(med, 3), round(p95, 3),
                   round(std, 3), len(durs), round(pct, 2)])
    ws.append(["TOTAL", "", round(total_avg, 3), "", "", "", len(per_step_paths), 100.0])
    out_wb.save(output_xlsx)
    print(f"  Aggregated {len(pooled)} kernels across {len(per_step_paths)} steps → {output_xlsx}")

    # Sister CSV for quick inspection / cross-platform diff scripts
    csv_path = output_xlsx[:-5] + ".csv" if output_xlsx.endswith(".xlsx") else output_xlsx + ".csv"
    import csv as _csv
    with open(csv_path, "w", newline="") as f:
        w = _csv.writer(f)
        w.writerow(["cpu_module", "gpu_kernel", "avg_us", "median_us", "p95_us",
                    "std_us", "n_steps", "pct%"])
        for key in order:
            durs = pooled[key]
            durs_sorted = sorted(durs)
            avg = means[key]
            med = percentile(durs_sorted, 50)
            p95 = percentile(durs_sorted, 95)
            std = pstdev(durs) if len(durs) > 1 else 0.0
            pct = 100 * avg / total_avg if total_avg > 0 else 0
            w.writerow([key[0], key[1],
                        f"{avg:.3f}", f"{med:.3f}", f"{p95:.3f}",
                        f"{std:.3f}", len(durs), f"{pct:.2f}"])
    print(f"  Companion CSV: {csv_path}")


def detect_multi_compiled_graph(capture_events, cg_event):
    """Detect if capture_graph contains multiple CompiledFxGraph blocks.

    Returns list of CompiledFxGraph events if multi-EP, empty list if standard.
    """
    cg_start = cg_event["ts"]
    cg_end = cg_start + cg_event.get("dur", 0)

    compiled_graphs = []
    for e in capture_events:
        if e.get("ph") != "X":
            continue
        name = e.get("name", "")
        if "CompiledFxGraph" not in name:
            continue
        e_start = e.get("ts", 0)
        e_end = e_start + e.get("dur", 0)
        if e_start >= cg_start and e_end <= cg_end:
            compiled_graphs.append(e)

    return compiled_graphs


def flatten_capture_events_for_multi_ep(capture_events, cg_event, compiled_graphs):
    """Flatten multi-CompiledFxGraph hierarchy for parse_trace compatibility.

    parse_trace expects: capture_graph → module → kernel_launch (2 levels)
    Multi-EP has:       capture_graph → CompiledFxGraph → module → kernel_launch (3 levels)

    Fix: remove CompiledFxGraph wrapper events from capture_events, so their
    children become direct children of capture_graph. We do this by shrinking
    each CompiledFxGraph event's duration to 0 (making it invisible to
    EventIndex.get_direct_children which checks time containment).
    """
    # Build set of CompiledFxGraph event IDs for fast lookup
    cg_ids = set(id(e) for e in compiled_graphs)

    flattened = []
    for e in capture_events:
        if id(e) in cg_ids:
            # Shrink CompiledFxGraph to 0 duration so its children
            # "escape" and become direct children of capture_graph
            e_copy = dict(e)
            e_copy["dur"] = 0
            flattened.append(e_copy)
        else:
            flattened.append(e)

    return flattened


def find_capture_graph_event(capture_events, target_bs):
    """Find the capture_graph_bs_N event matching target_bs."""
    target_name = f"capture_graph_bs_{target_bs}"

    # Exact match
    matches = [
        e for e in capture_events
        if e.get("name") == target_name and e.get("ph") == "X"
        and e.get("dur", 0) > 100  # filter out the tiny duplicate events
    ]
    if matches:
        return matches[0]

    # Fallback: largest bs smaller than target
    bs_events = []
    for e in capture_events:
        if e.get("ph") != "X":
            continue
        m = re.match(r"^capture_graph_bs_(\d+)$", e.get("name", ""))
        if m and e.get("dur", 0) > 100:
            bs_events.append((int(m.group(1)), e))

    if bs_events:
        # Try closest smaller
        smaller = [(b, e) for b, e in bs_events if b <= target_bs]
        if smaller:
            best = max(smaller, key=lambda x: x[0])
            return best[1]
        # Any
        return bs_events[0][1]

    return None


def main():
    parser = argparse.ArgumentParser(
        description="parse_trace.py wrapper with --target-bs and multi-EP support."
    )
    parser.add_argument("filepath", help="Path to trace JSON or JSON.GZ file")
    parser.add_argument("--layer", type=int, default=3, help="Target layer index")
    parser.add_argument(
        "--target-bs", type=int, default=None,
        help="Decode batch size (default: most frequent)",
    )
    parser.add_argument(
        "--skip-warmup", type=int, default=5,
        help="Drop first N decode steps (cold caches / JIT). Matches B300/B200 "
             "trace_layer_detail.py convention. Default 5.",
    )
    parser.add_argument(
        "--max-steps", type=int, default=20,
        help="Aggregate over N steady-state decodes (default 20 → ~600 samples "
             "per kernel). 0 = all available. Set 1 + --skip-ratio for old "
             "single-decode behavior.",
    )
    parser.add_argument(
        "--skip-ratio", type=float, default=None,
        help="DEPRECATED: position-based picker (0.0=first, 0.5=median). "
             "Only honored when --max-steps=1. Otherwise --skip-warmup wins.",
    )
    parser.add_argument(
        "--capture-trace", type=str, default=None,
        help="Explicit path to capture_graph trace file (bypasses auto-detection)",
    )
    parser.add_argument(
        "--suffix", type=str, default=None,
        help="Suffix for output files (e.g., 'c64'). Auto-detected from filename if not set.",
    )
    args = parser.parse_args()

    filepath = args.filepath

    # Auto-detect suffix from filename (e.g., ..._c64_full.log → c64)
    if args.suffix is None:
        m = re.search(r'_(c\d+)[_.]', os.path.basename(filepath))
        args.suffix = m.group(1) if m else None

    print(f"Loading run trace: {filepath}")
    trace = parse_trace.load_trace(filepath)
    events = trace.get("traceEvents", [])
    print(f"Loaded {len(events)} events")

    if args.capture_trace:
        capture_trace_path = args.capture_trace
        print(f"Using explicit capture trace: {capture_trace_path}")
    else:
        capture_trace_path = parse_trace.find_capture_graph_trace_path(filepath)
    if capture_trace_path is None:
        print("Warning: no capture trace found, using run trace for hierarchy")
        capture_events = events
    else:
        print(f"Loading capture trace: {capture_trace_path}")
        capture_events = parse_trace.load_trace(capture_trace_path).get(
            "traceEvents", []
        )
        print(f"Loaded {len(capture_events)} capture events")

    # --- Prefill: unchanged ---
    print("\n" + "=" * 60)
    print("PREFILL ANALYSIS")
    print("=" * 60)
    pfx = f"_{args.suffix}" if args.suffix else ""
    parse_trace.parse_prefill(events, f"prefill_breakdown{pfx}.xlsx", target_layer=args.layer)

    # --- Decode: select target bs ---
    print("\n" + "=" * 60)
    print("DECODE ANALYSIS (with --target-bs and multi-EP support)")
    print("=" * 60)

    # R8d: select N steady-state decodes (default skip 5, take 20).
    # Back-compat: pass --max-steps 1 + --skip-ratio for old single-decode behavior.
    target_decodes, actual_bs = select_decodes_bs(
        events, args.target_bs,
        skip_warmup=args.skip_warmup, max_steps=args.max_steps,
        skip_ratio=args.skip_ratio,
    )
    if not target_decodes:
        print("ERROR: no decode events at target bs")
        sys.exit(1)

    print(f"Will aggregate {len(target_decodes)} decode steps")

    # --- Multi-EP detection and flattening (computed once, reused per step) ---
    cg_event = find_capture_graph_event(capture_events, actual_bs)
    if cg_event is None:
        print("WARNING: No capture_graph event found, proceeding without multi-EP fix")
        final_capture_events = capture_events
    else:
        print(f"Using capture_graph: {cg_event.get('name')} (dur={cg_event.get('dur',0)}µs)")
        compiled_graphs = detect_multi_compiled_graph(capture_events, cg_event)

        if len(compiled_graphs) > 1:
            print(f"\n*** MULTI-EP DETECTED: {len(compiled_graphs)} CompiledFxGraph blocks ***")
            for i, cg in enumerate(compiled_graphs):
                print(f"  [{i}] {cg.get('name', '')[:60]}... (dur={cg.get('dur',0)}µs)")
            print("Flattening hierarchy for parse_trace compatibility...")
            final_capture_events = flatten_capture_events_for_multi_ep(
                capture_events, cg_event, compiled_graphs
            )
            print(f"Flattened: {len(capture_events)} → {len(final_capture_events)} events")
        else:
            print("Standard single-graph hierarchy (EP=1 or single CompiledFxGraph)")
            final_capture_events = capture_events

    # --- Per-decode parse + aggregate (R8d wide-sample) ---
    # parse_trace.parse_decode wants exactly ONE decode in its event list, so we
    # run it per-step then merge the N per-step xlsx files via merge_decode_xlsx.
    final_xlsx = f"decode_breakdown{pfx}.xlsx"
    if len(target_decodes) == 1:
        d = target_decodes[0]
        filtered = [e for e in events
                    if not (e.get("name", "").startswith("decode[")
                            and e.get("ph") == "X"
                            and e.get("cat") == "gpu_user_annotation"
                            and e is not d)]
        parse_trace.parse_decode(filtered, final_capture_events, final_xlsx,
                                  target_layer=args.layer)
        print(f"Single-decode xlsx: {final_xlsx}")
    else:
        per_step_paths = []
        per_step_dir = f"decode_per_step{pfx}"
        os.makedirs(per_step_dir, exist_ok=True)
        for i, d in enumerate(target_decodes):
            tmp = os.path.join(per_step_dir, f"step_{i:02d}.xlsx")
            filtered = [e for e in events
                        if not (e.get("name", "").startswith("decode[")
                                and e.get("ph") == "X"
                                and e.get("cat") == "gpu_user_annotation"
                                and e is not d)]
            try:
                import io, contextlib
                cap = io.StringIO()
                with contextlib.redirect_stdout(cap):
                    parse_trace.parse_decode(filtered, final_capture_events, tmp,
                                              target_layer=args.layer)
                if os.path.exists(tmp):
                    per_step_paths.append(tmp)
                    if i < 3 or i == len(target_decodes) - 1:
                        print(f"  step {i:02d}: parse_decode OK -> {tmp}")
                else:
                    # Silent skip: no xlsx written but no exception either.
                    # Surface captured stdout so downstream debugging knows why
                    # (typical reasons: 'No norm module', 'no decode events',
                    # capture_graph hierarchy mismatch — all printed by ATOM).
                    captured = cap.getvalue().strip().splitlines()
                    tail = "\n".join("    | " + l for l in captured[-12:]) if captured else "    (no captured stdout)"
                    print(f"  step {i:02d}: parse_decode silently skipped — captured stdout tail:")
                    print(tail)
            except SystemExit as e:
                captured = cap.getvalue().strip().splitlines()
                tail = "\n".join("    | " + l for l in captured[-8:]) if captured else ""
                print(f"  step {i:02d}: parse_decode SystemExit code={e.code}; captured tail:")
                if tail: print(tail)
            except Exception as e:
                captured = cap.getvalue().strip().splitlines()
                tail = "\n".join("    | " + l for l in captured[-8:]) if captured else ""
                print(f"  step {i:02d}: parse_decode exception: {e!r}; captured tail:")
                if tail: print(tail)
        if per_step_paths:
            print(f"\nMerging {len(per_step_paths)}/{len(target_decodes)} step xlsx files...")
            merge_decode_xlsx(per_step_paths, final_xlsx)
        else:
            # All steps silently skipped — typically 'No norm module found in
            # capture_graph modules' on ROCm 7.2.2+ ATOM. parse_trace.parse_decode
            # requires the older norm-module hierarchy that newer ATOM stopped
            # annotating. Fall back to decode_kernel_breakdown.py which uses
            # mla_a8w8 kernel anchor — doesn't need norm modules.
            print(f"\nAll {len(target_decodes)} steps silently skipped by parse_trace.parse_decode "
                  f"(typical: 'No norm module' on ROCm 7.2.2+).")
            print(f"Falling back to scripts/decode_kernel_breakdown.py (kernel-anchor splitting)...")
            import subprocess
            fallback_script = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                            "decode_kernel_breakdown.py")
            if not os.path.exists(fallback_script):
                print(f"  ERROR: fallback script not found at {fallback_script}")
                sys.exit(1)
            cmd = ["python3", fallback_script, args.filepath,
                   "--target-bs", str(actual_bs),
                   "--skip-warmup", str(args.skip_warmup),
                   "--max-steps", str(args.max_steps),
                   "--output", final_xlsx]
            if args.suffix:
                cmd.extend(["--layers", "10-40"])  # decode_kernel_breakdown uses --layers, default 10-40
            print(f"  $ {' '.join(cmd)}")
            try:
                result = subprocess.run(cmd, capture_output=True, text=True, check=False)
                # Stream child output so user sees progress
                if result.stdout: print(result.stdout)
                if result.stderr: print(result.stderr, file=sys.stderr)
                if result.returncode != 0:
                    print(f"  FALLBACK FAILED with exit code {result.returncode}")
                    sys.exit(result.returncode)
                if os.path.exists(final_xlsx):
                    print(f"  Fallback xlsx written: {final_xlsx}")
                else:
                    print(f"  ERROR: fallback completed but {final_xlsx} not produced")
                    sys.exit(1)
            except FileNotFoundError as e:
                print(f"  Fallback subprocess failed: {e}")
                sys.exit(1)


if __name__ == "__main__":
    main()
