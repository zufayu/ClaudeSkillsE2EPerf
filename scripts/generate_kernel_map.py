#!/usr/bin/env python3
"""
Generate B200 vs MI355X kernel map CSV from:
  - B200 layer_kernel_avg.csv (from trace_layer_detail.py)
  - MI355X decode_breakdown.xlsx (from run_parse_trace.py)

Output: kernel_map CSV with strict timeline alignment, overlap, and PASS grouping.

Rules:
  1. No missing operators — every B200 and MI355X kernel appears
  2. Original kernel names preserved exactly
  3. Data sources noted in footer
  4. Sum verification: total == sum of individual values

Usage:
    python3 generate_kernel_map.py \\
        --b200-csv layer_kernel_avg.csv \\
        --mi355x-xlsx decode_breakdown.xlsx \\
        --output b200_vs_mi355x_kernel_map.csv
"""

import argparse
import csv
import os
import sys

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ── B200 operator → PASS classification ──
# PASS_MAP — direct operator → PASS lookup (user's 7-PASS taxonomy).
# Used by get_pass() for canonical operator names; pattern-matching fallback
# for 'other:' prefixed kernels handled in get_pass() body.
PASS_MAP = {
    "EP_AR+residual+RMSNorm(fused)": "EP_AR",
    "qkv_a_proj_GEMM":                  "MHA",
    "qkv_a_splitK_reduce":              "MHA",
    "q/k_norm_RMSNorm":                 "MHA",
    "q_b_proj_GEMM":                    "MHA",
    "uk_gemm(K_expansion)":             "MHA",
    "k_concat":                         "MHA",
    "RoPE+KV_cache_write":              "MHA",
    "set_mla_kv":                       "MHA",
    "Attention(FMHA)":                  "MHA",
    "uv_gemm(V_expansion)":             "MHA",
    "q_b_proj_GEMM #2":                 "MHA",
    "o_proj_quant(BF16→FP4)":           "O_proj",
    "o_proj_GEMM":                      "O_proj",
    "o_proj_GEMM #1":                   "O_proj",
    "o_proj_GEMM #2":                   "O_proj",
    "o_proj_quant(BF16→FP4) #1":        "O_proj",
    "o_proj_quant(BF16→FP4) #2":        "O_proj",
    "router_GEMM":                      "MoE_Route",
    "router_splitK_reduce":             "MoE_Route",
    "MoE_input_quant(BF16→FP4)":        "MoE_Route",
    "TopK_select":                      "MoE_Route",
    "expert_sort":                      "MoE_Route",
    "Moe_Expert_quant(BF16→FP4)":       "MoE_Expert",
    "gate_up_GEMM(+SwiGLU)":            "MoE_Expert",
    "down_GEMM":                        "MoE_Expert",
    "MoE_finalize+residual":            "MoE_Expert",
    "MoE_finalize":                     "MoE_Expert",
    "shared_quant(BF16→FP4)":           "Shared_Exp",
    "shared_GEMM(FP4)":                 "Shared_Exp",
    "SiLU×Mul":                         "Shared_Exp",
    "tensor_copy":                      "Residual",
    "residual_add":                     "Residual",
}

# ── MI355X module → PASS classification ──
# Used to classify MI355X kernels independently of which B200 row they're mapped to.
# This is needed because B200's cross-stream overlap can place o_proj_GEMM after
# router_GEMM, causing MI355X MoE kernels to be mapped to a B200 O_proj row.
MI355X_MODULE_PASS = {
    "input_layernorm":                          "EP_AR",
    "gemm_a16w16":                              None,   # position-dependent (MHA or MoE_Route)
    "hipLaunchKernel":                          "MHA",
    "q_proj_and_k_up_proj":                     "MHA",
    "rope_and_kv_cache":                        "MHA",
    "mla_decode":                               "MHA",
    "v_up_proj_and_o_proj":                     "MHA",
    "triton_poi_fused_as_strided_clone_copy__0":"MHA",
    "post_attn_layernorm":                      "EP_AR",
    "triton_poi_fused_as_strided_clone_1":      "EP_AR",
    "triton_poi":                               "EP_AR",
    "rocm_aiter_biased_grouped_topk_impl":      "MoE_Route",
    "mxfp4_moe":                                None,   # position-dependent: kernel-name disambiguates router/expert/shared
}


def get_pass(operator, occurrence_num):
    """Get PASS name for a B200 operator.

    Operators from trace_layer_detail.py may have 'other:' prefix with full
    kernel name (e.g. 'other:void router_gemm_kernel_float_output<...').
    Pattern-match these to the correct PASS using user's 7-PASS taxonomy.
    """
    p = PASS_MAP.get(operator)
    if p is not None:
        return p
    # EP_AR — both occurrences map to the same PASS (no longer split into
    # before_MHA / before_MOE; user's taxonomy treats them uniformly)
    if "EP_AR" in operator:
        return "EP_AR"
    # Pattern-match 'other:' prefixed operators by kernel name keywords
    op_lower = operator.lower()
    if "router_gemm" in op_lower or "routingmainkernel" in op_lower or "routingindices" in op_lower:
        return "MoE_Route"
    if "fused_a_gemm" in op_lower:
        return "MoE_Route"
    if "act_and_mul" in op_lower or "shared" in op_lower:
        return "Shared_Exp"
    if "finalizekernel" in op_lower or "finalize" in op_lower:
        return "MoE_Expert"
    if "bmm_" in op_lower and ("e2m1" in op_lower or "bfloat16" in op_lower):
        return "MoE_Expert"
    # other/unknown
    return "Other"


def get_mi355x_pass(module, b200_pass):
    """Get PASS for MI355X kernel, using its module name when available.

    Falls back to b200_pass when the MI355X module is unknown or position-dependent.
    """
    if not module:
        return b200_pass
    p = MI355X_MODULE_PASS.get(module)
    if p is not None:
        return p
    # For position-dependent modules (gemm_a16w16), use B200 PASS
    return b200_pass


def parse_b200_csv(path):
    """Parse B200 layer_kernel_avg.csv. Returns list of dicts."""
    rows = []
    totals = {}
    with open(path, "r") as f:
        reader = csv.reader(f)
        header = next(reader)

        # Detect column positions by name — B200 columns only (exclude MI355X)
        col_map = {}
        for i, h in enumerate(header):
            h_raw = h.strip()
            h_lower = h_raw.lower()
            if "mi355x" in h_lower:
                continue  # skip MI355X columns
            if "operator" in h_lower:
                col_map["operator"] = i
            elif "raw_kernel" in h_lower:
                col_map["raw_kernel"] = i
            elif "stream" in h_lower:
                col_map["stream"] = i
            elif "avg_us" in h_lower and "overlap" not in h_lower:
                col_map["avg_us"] = i
            elif "overlap_us" in h_lower:
                col_map["overlap_us"] = i
            elif "overlap_with" in h_lower:
                col_map["overlap_with"] = i
            elif "module" in h_lower:
                col_map["module"] = i
            elif h_raw == "#":
                col_map["num"] = i

        for row in reader:
            if not row or all(c.strip() == "" for c in row):
                continue
            # Summary rows
            joined = ",".join(row)
            if "TOTAL" in joined or "Walltime" in joined or "Overlap" in joined or "PASS" in joined or "Silicon" in joined:
                # Parse totals
                for i, c in enumerate(row):
                    c = c.strip()
                    if "kernel_sum" in c or "TOTAL" in c:
                        for j in range(i + 1, len(row)):
                            try:
                                totals["kernel_sum"] = float(row[j].strip())
                                break
                            except (ValueError, IndexError):
                                continue
                    if "Walltime" in c:
                        for j in range(i + 1, len(row)):
                            try:
                                totals["walltime"] = float(row[j].strip())
                                break
                            except (ValueError, IndexError):
                                continue
                    if c == "Overlap" or "B200 Overlap" in c:
                        for j in range(i + 1, len(row)):
                            try:
                                totals["overlap"] = float(row[j].strip())
                                break
                            except (ValueError, IndexError):
                                continue
                continue

            operator = row[col_map["operator"]].strip() if "operator" in col_map else ""
            if not operator:
                continue

            try:
                avg_us = float(row[col_map["avg_us"]].strip()) if "avg_us" in col_map else 0
            except (ValueError, IndexError):
                continue

            overlap_us = 0
            if "overlap_us" in col_map:
                try:
                    overlap_us = float(row[col_map["overlap_us"]].strip())
                except (ValueError, IndexError):
                    pass

            rows.append({
                "operator": operator,
                "raw_kernel": row[col_map.get("raw_kernel", 0)].strip() if "raw_kernel" in col_map else "",
                "stream": row[col_map.get("stream", 0)].strip() if "stream" in col_map else "",
                "avg_us": avg_us,
                "overlap_us": overlap_us,
                "overlap_with": row[col_map.get("overlap_with", 0)].strip() if "overlap_with" in col_map else "",
                "module": row[col_map.get("module", 0)].strip() if "module" in col_map else "",
            })

    return rows, totals


def parse_mi355x_xlsx(path):
    """Parse MI355X decode_breakdown.xlsx. Returns list of dicts."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # header: cpu_module, gpu_kernel, duration_us, pct%, sum_per_module, module_pct%, avg_duration_us, avg_sum_per_module
    rows = []
    total_avg = None
    for r in all_rows[1:]:
        module = r[0] or ""
        kernel = r[1] or ""
        avg_us = r[6]  # avg duration_us

        if str(module).strip() == "TOTAL":
            total_avg = float(avg_us) if avg_us else None
            continue

        if avg_us is None:
            continue

        rows.append({
            "module": str(module).strip(),
            "kernel": str(kernel).strip(),
            "avg_us": float(avg_us),
        })

    # Fill in parent_module: the module each row actually belongs to
    # (inheriting from previous row when module is empty)
    cur_mod = ""
    for r in rows:
        if r["module"]:
            cur_mod = r["module"]
        r["parent_module"] = cur_mod

    return rows, total_avg


def classify_b200_op(operator):
    """Map B200 operator name to a logical operator group."""
    op = operator.lower()
    # Order matters: check specific patterns before general ones
    if "ep_ar" in op:
        return "comm"  # disambiguated later by occurrence count
    if "router" in op:
        return "moe_router"
    if "qkv_a" in op or "splitk_reduce" in op:
        return "mla_qkv_a"
    if "q/k_norm" in op or "q_norm" in op or "k_norm" in op:
        return "mla_qk_norm"
    if "q_b_proj" in op:
        return "mla_q_b_k_up"
    if "uk_gemm" in op or "k_concat" in op:
        return "mla_q_b_k_up"
    if "rope" in op or "set_mla_kv" in op:
        return "mla_rope_cache"
    if "attention" in op or "fmha" in op:
        return "mla_attn"
    if "uv_gemm" in op:
        return "mla_uv_o_proj"
    if "o_proj" in op:
        return "mla_uv_o_proj"
    if "router" in op:
        return "moe_router"
    # MoE_Route: routing GEMM, TopK, expert sort, MoE input quant
    if any(kw in op for kw in ["topk", "expert_sort", "moe_input_quant",
                                "routingmain", "routingindices", "fused_a_gemm"]):
        return "moe_router"
    # Shared_Exp: shared expert (silu/mul, shared_quant, shared_GEMM)
    if any(kw in op for kw in ["silu", "shared", "act_and_mul"]):
        return "shared_compute"
    # Residual: tensor_copy, residual_add
    if any(kw in op for kw in ["tensor_copy", "residual_add"]):
        return "residual"
    # MoE_Expert: gate_up_GEMM, down_GEMM, MoE_finalize
    if any(kw in op for kw in ["gate_up", "down_gemm", "moe_finalize", "moe_expert_quant",
                                "bmm_", "finalizekernel"]):
        return "moe_expert"
    return "other"


def classify_mi355x_row(module, kernel, gemm_a16w16_count):
    """Map MI355X module+kernel to a logical operator group.

    gemm_a16w16_count tracks how many gemm_a16w16 modules we've seen
    (1st = qkv_a, 2nd = router).

    Kernel name is checked FIRST for certain patterns that override module
    classification (e.g. norm kernels inside q_proj_and_k_up_proj module).
    """
    kn = kernel.lower() if kernel else ""
    mod = module.lower() if module else ""

    # ── Kernel-name overrides (priority over module) ──
    # Norm kernels can appear inside q_proj_and_k_up_proj or hipLaunchKernel
    if "fused_qk_rmsnorm" in kn or "add_rmsnorm_quant" in kn or "rmsnorm_quant" in kn:
        return "mla_qk_norm"
    # MLA reduce can appear as sub-kernel under mla_decode
    if "kn_mla_reduce" in kn:
        return "mla_attn"

    # ── Module-based classification ──
    if "input_layernorm" in mod:
        return "comm_pre_attn"
    if "per_token_quant" in mod:
        return "mla_input_quant"  # disambiguated at call site for 2nd occurrence
    if "gemm_a8w8_bpreshuffle" in mod:
        return "mla_qkv_a"
    if "gemm_a16w16" in mod:
        return "mla_qkv_a" if gemm_a16w16_count <= 1 else "moe_router"
    if "hiplaunchkernel" in mod:
        return "mla_qk_norm"
    if "q_proj_and_k_up_proj" in mod:
        return "mla_q_b_k_up"
    if "rope_and_kv_cache" in mod:
        return "mla_rope_cache"
    if "mla_decode" in mod:
        return "mla_attn"
    if "v_up_proj_and_o_proj" in mod:
        return "mla_uv_o_proj"
    if "triton_poi_fused_as_strided_clone_copy__0" in mod:
        return "mla_transpose"
    if "post_attn_layernorm" in mod:
        return "comm_pre_moe"
    if "triton_poi_fused_as_strided_clone_1" in mod:
        return "comm_pre_moe"
    if "triton_poi" in mod:
        return "comm_pre_moe"
    if "rocm_aiter_biased_grouped_topk" in mod:
        return "moe_router"
    if "mxfp4_moe" in mod:
        # mxfp4_moe lumps router (sort/topk) + expert (gemm) + shared (quant) on MI355X.
        # Order matters: `mxfp4_quant_moe_sort_kernel` matches BOTH 'quant' and 'sort'
        # — it's fundamentally a shared-expert quant (per user's manual alignment table),
        # so check 'mxfp4_quant' BEFORE generic 'sort'.
        if "moe_gemm" in kn or "moe_mxgemm" in kn:
            return "moe_expert"
        if "mxfp4_quant" in kn:
            return "shared_compute"   # quant kernel for shared expert (B200 cvt_fp16_to_fp4)
        if "grouped_topk" in kn or "topk_opt_sort" in kn or "biased_grouped_topk" in kn:
            return "moe_router"        # explicit topk → router (TopK_select on B200)
        if "moesorting" in kn or "moesort" in kn:
            return "moe_router"        # expert sort phases → router (expert_sort on B200)
        return "moe_router"            # safe default for unrecognized mxfp4_moe phases

    # ── Fallback: kernel name only (when module is empty or unknown) ──
    if "reduce_scatter" in kn or "local_device_load_rmsnorm" in kn:
        return "comm_pre_attn"
    if "gemm_xdl" in kn or "cijk_" in kn:
        return "mla_q_b_k_up"
    if "batched_gemm_a8w8" in kn:
        return "mla_q_b_k_up"
    if "flatmm" in kn:
        return "mla_uv_o_proj"
    if "mla_a8w8" in kn:
        return "mla_attn"
    if "fuse_qk_rope" in kn:
        return "mla_rope_cache"
    if "bf16gemm" in kn:
        return "mla_qkv_a"
    if "kernel_moe_mxgemm" in kn or "moe_gemm" in kn:
        return "moe_expert"
    if "moesorting" in kn or "grouped_topk" in kn:
        return "moe_router"
    if "mxfp4_quant" in kn:
        return "shared_compute"
    if "per_token_scaled_quant" in kn:
        return "mla_input_quant"

    return "other"


# Canonical order of logical operator groups (execution order in a layer)
LOGICAL_OP_ORDER = [
    "comm_pre_attn",
    "mla_input_quant",
    "mla_qkv_a",
    "mla_qk_norm",
    "mla_q_b_k_up",
    "mla_rope_cache",
    "mla_attn",
    "mla_uv_o_proj",
    "mla_transpose",
    "comm_pre_moe",
    "moe_router",
    "shared_compute",
    "moe_expert",
    "residual",
    "other",
]

# Map logical operator group → PASS name (user's 7-PASS taxonomy)
LOGOP_TO_PASS = {
    "comm_pre_attn":   "EP_AR",
    "mla_input_quant": "MHA",
    "mla_qkv_a":       "MHA",
    "mla_qk_norm":     "MHA",
    "mla_q_b_k_up":    "MHA",
    "mla_rope_cache":  "MHA",
    "mla_attn":        "MHA",
    "mla_uv_o_proj":   "O_proj",
    "mla_transpose":   "O_proj",
    "comm_pre_moe":    "EP_AR",
    "moe_router":      "MoE_Route",
    "shared_compute":  "Shared_Exp",
    "moe_expert":      "MoE_Expert",
    "residual":        "Residual",
    "other":           "Other",
}


def group_by_logop(rows, classify_fn):
    """Group rows into (logical_op, [rows]) preserving order within groups.

    Returns list of (logop, [row_dicts]) in the order groups first appear.
    """
    groups = {}
    order = []
    for row in rows:
        logop = classify_fn(row)
        if logop not in groups:
            groups[logop] = []
            order.append(logop)
        groups[logop].append(row)
    return [(logop, groups[logop]) for logop in order]


def generate_map(b200_rows, b200_totals, mi355x_rows, mi355x_total, b200_csv_path, mi355x_xlsx_path, output_path):
    """Generate kernel map CSV using semantic operator alignment.

    Instead of aligning by position (which breaks with B200 dual-stream overlap),
    we classify both sides into logical operator groups, then align groups.
    """

    # ── Pass 1: Classify B200 rows ──
    ep_ar_count = 0
    b200_classified = []
    for b in b200_rows:
        op = b["operator"]
        logop = classify_b200_op(op)
        # Disambiguate comm (EP_AR) by occurrence
        if logop == "comm":
            if "#2" in op:
                logop = "comm_pre_moe"
            elif "#1" in op:
                logop = "comm_pre_attn"
            else:
                ep_ar_count += 1
                logop = "comm_pre_attn" if ep_ar_count == 1 else "comm_pre_moe"
        b200_classified.append((logop, b))

    # ── Pass 1: Classify MI355X rows ──
    # Check if FP8 model: if gemm_a8w8_bpreshuffle exists, qkv_a is FP8
    # and gemm_a16w16 is always router (not qkv_a)
    has_fp8_qkv_a = any(
        "gemm_a8w8_bpreshuffle" in (mi["module"] or mi.get("parent_module", "")).lower()
        for mi in mi355x_rows
    )
    gemm_a16w16_count = 0
    per_token_quant_count = 0
    mi355x_classified = []
    for mi in mi355x_rows:
        mod = mi["module"] or mi.get("parent_module", "")
        mod_lower = mod.lower()
        if "gemm_a16w16" in mod_lower and mi["module"]:
            gemm_a16w16_count += 1
        if "per_token_quant" in mod_lower and mi["module"]:
            per_token_quant_count += 1
        # If FP8 model has gemm_a8w8_bpreshuffle for qkv_a,
        # then gemm_a16w16 is always router regardless of occurrence count
        effective_count = gemm_a16w16_count
        if has_fp8_qkv_a and "gemm_a16w16" in mod_lower:
            effective_count = 2  # force router classification
        logop = classify_mi355x_row(mod, mi["kernel"], effective_count)
        # Fix: 2nd per_token_quant is before o_proj, not input quant
        if logop == "mla_input_quant" and per_token_quant_count > 1:
            logop = "mla_uv_o_proj"
        mi355x_classified.append((logop, mi))

    # ── Pass 2: Group by logical operator ──
    b200_groups = group_by_logop(b200_classified, lambda x: x[0])
    b200_groups = [(logop, [item[1] for item in items]) for logop, items in b200_groups]

    mi355x_groups = group_by_logop(mi355x_classified, lambda x: x[0])
    mi355x_groups = [(logop, [item[1] for item in items]) for logop, items in mi355x_groups]

    # Build lookup dicts
    b200_by_logop = {logop: rows for logop, rows in b200_groups}
    mi355x_by_logop = {logop: rows for logop, rows in mi355x_groups}

    # Determine full ordered list of logical ops (union of both sides)
    all_logops = []
    for logop in LOGICAL_OP_ORDER:
        if logop in b200_by_logop or logop in mi355x_by_logop:
            all_logops.append(logop)
    # Add any that aren't in the canonical order
    for logop in list(b200_by_logop.keys()) + list(mi355x_by_logop.keys()):
        if logop not in all_logops:
            all_logops.append(logop)

    # ── Pass 3: Align within each logical operator group ──
    output_rows = []
    pass_b200 = {}
    pass_mi355x = {}
    pass_nv_kernels = {}
    pass_amd_kernels = {}
    b200_sum_check = 0.0
    mi355x_sum_check = 0.0
    unmapped_mi355x = []

    for logop in all_logops:
        b200_list = b200_by_logop.get(logop, [])
        mi355x_list = mi355x_by_logop.get(logop, [])
        pass_name = LOGOP_TO_PASS.get(logop, "other")

        # Pair rows within the group by position
        max_len = max(len(b200_list), len(mi355x_list))

        for i in range(max_len):
            b200 = b200_list[i] if i < len(b200_list) else None
            mi = mi355x_list[i] if i < len(mi355x_list) else None

            row = {
                "b200_operator": "",
                "b200_raw_kernel": "",
                "b200_stream": "",
                "b200_avg_us": "",
                "b200_overlap_us": "",
                "b200_overlap_with": "",
                "mi355x_module": "",
                "mi355x_kernel": "",
                "mi355x_avg_us": "",
                "notes": "",
                "pass": pass_name,
            }

            if b200:
                row["b200_operator"] = b200["operator"]
                row["b200_raw_kernel"] = b200["raw_kernel"]
                row["b200_stream"] = b200["stream"]
                row["b200_avg_us"] = b200["avg_us"]
                row["b200_overlap_us"] = b200["overlap_us"]
                row["b200_overlap_with"] = b200["overlap_with"]
                b200_sum_check += b200["avg_us"]
                pass_b200[pass_name] = pass_b200.get(pass_name, 0) + b200["avg_us"]
                if pass_name not in pass_nv_kernels:
                    pass_nv_kernels[pass_name] = []
                pass_nv_kernels[pass_name].append(b200["raw_kernel"])

            if mi:
                row["mi355x_module"] = mi["module"]
                row["mi355x_kernel"] = mi["kernel"]
                row["mi355x_avg_us"] = mi["avg_us"]
                mi355x_sum_check += mi["avg_us"]
                mi_pass = get_mi355x_pass(mi["parent_module"], pass_name)
                pass_mi355x[mi_pass] = pass_mi355x.get(mi_pass, 0) + mi["avg_us"]
                if mi_pass not in pass_amd_kernels:
                    pass_amd_kernels[mi_pass] = []
                pass_amd_kernels[mi_pass].append(mi["kernel"])

            # Notes for unmatched rows
            if b200 and not mi:
                row["notes"] = "B200 only"
            elif mi and not b200:
                row["notes"] = f"MI355X only ({logop})"
                unmapped_mi355x.append(mi)

            output_rows.append(row)

    # ── Checksum verification ──
    b200_expected = b200_totals.get("kernel_sum")
    mi355x_expected = mi355x_total

    checksum_notes = []
    if b200_expected is not None:
        diff = abs(b200_sum_check - b200_expected)
        if diff > 0.5:
            checksum_notes.append(f"WARNING: B200 sum mismatch: computed={b200_sum_check:.2f} expected={b200_expected:.1f} diff={diff:.2f}")
        else:
            checksum_notes.append(f"OK: B200 sum={b200_sum_check:.2f} matches expected={b200_expected:.1f}")

    if mi355x_expected is not None:
        diff = abs(mi355x_sum_check - mi355x_expected)
        if diff > 0.2:
            checksum_notes.append(f"WARNING: MI355X sum mismatch: computed={mi355x_sum_check:.2f} expected={mi355x_expected:.2f} diff={diff:.2f}")
        else:
            checksum_notes.append(f"OK: MI355X sum={mi355x_sum_check:.2f} matches expected={mi355x_expected:.2f}")

    if unmapped_mi355x:
        checksum_notes.append(f"WARNING: {len(unmapped_mi355x)} MI355X operators unmapped")

    # ── Write XLSX ──
    # 11 columns total: PASS + 9 original data columns + Notes
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "kernel_map"

    def fmt(v, prec=1):
        if isinstance(v, (int, float)) and v != "":
            return f"{v:.{prec}f}"
        return v

    # Header (11 cols)
    ws.append([
        "PASS",
        "B200_Operator", "B200_Raw_Kernel", "B200_Stream", "B200_Avg_us",
        "B200_Overlap_us", "B200_Overlap_With",
        "MI355X_Module", "MI355X_Kernel", "MI355X_Avg_us",
        "Notes",
    ])

    # Data rows
    for r in output_rows:
        ws.append([
            r.get("pass", ""),
            r["b200_operator"], r["b200_raw_kernel"], r["b200_stream"],
            fmt(r["b200_avg_us"], 1),
            fmt(r["b200_overlap_us"], 1),
            r["b200_overlap_with"],
            r["mi355x_module"], r["mi355x_kernel"],
            fmt(r["mi355x_avg_us"], 2),
            r["notes"],
        ])

    # Blank line
    ws.append([])

    # Totals
    ws.append(["", "B200 TOTAL (kernel_sum)", "", "", f"{b200_sum_check:.1f}", "", "",
                "MI355X TOTAL", "", f"{mi355x_sum_check:.2f}", ""])
    if b200_totals.get("walltime"):
        ws.append(["", "B200 Walltime", "", "", f'{b200_totals["walltime"]:.1f}', "", "", "", "", "", ""])
    if b200_totals.get("overlap"):
        overlap = b200_totals["overlap"]
        pct = overlap / b200_sum_check * 100 if b200_sum_check > 0 else 0
        ws.append(["", "B200 Overlap", "", "", f"{overlap:.1f}", "", "", "", "", "",
                   f"{overlap:.1f}us overlap = {pct:.1f}% of kernel sum"])

    ws.append([])

    # PASS summary (user's 7-PASS taxonomy)
    ws.append(["", "", "Silicon NV", "Silicon AMD", "", "", "", "", "", "", ""])
    ws.append(["PASS", "", "B200", "MI355X", "gap", "NV_Kernels", "AMD_Kernels", "", "", "", ""])
    pass_order = ["EP_AR", "MHA", "O_proj", "MoE_Route", "Shared_Exp", "MoE_Expert", "Residual", "Other"]
    for pname in pass_order:
        b200_val = pass_b200.get(pname, 0)
        mi355x_val = pass_mi355x.get(pname, 0)
        gap = mi355x_val - b200_val
        nv_k = "\n".join(pass_nv_kernels.get(pname, []))
        amd_k = "\n".join(pass_amd_kernels.get(pname, []))
        if b200_val > 0 or mi355x_val > 0:
            ws.append([pname, "", f"{b200_val:.1f}", f"{mi355x_val:.2f}",
                       f"{gap:+.1f}", nv_k, amd_k, "", "", "", ""])

    ws.append([])

    # Footer: data sources + checksum
    ws.append(["# Data Sources:"])
    ws.append([f"#   B200: {os.path.abspath(b200_csv_path)}"])
    ws.append([f"#   MI355X: {os.path.abspath(mi355x_xlsx_path)}"])
    ws.append(["# Checksum:"])
    for note in checksum_notes:
        ws.append([f"#   {note}"])

    wb.save(output_path)

    print(f"Written: {output_path}")
    print(f"  B200 operators: {len(b200_rows)}, MI355X operators: {len(mi355x_rows)}")
    print(f"  Output rows: {len(output_rows)}")
    for note in checksum_notes:
        print(f"  {note}")


def main():
    parser = argparse.ArgumentParser(description="Generate B200 vs MI355X kernel map")
    parser.add_argument("--b200-csv", required=True, help="B200 layer_kernel_avg.csv")
    parser.add_argument("--mi355x-xlsx", required=True, help="MI355X decode_breakdown.xlsx")
    parser.add_argument("--output", default="kernel_map_b200_vs_mi355x.xlsx",
                        help="Output XLSX path (default: kernel_map_b200_vs_mi355x.xlsx)")
    args = parser.parse_args()
    if not args.output.endswith(".xlsx"):
        print(f"WARNING: --output '{args.output}' doesn't end in .xlsx; this script writes XLSX format only.", file=sys.stderr)

    b200_rows, b200_totals = parse_b200_csv(args.b200_csv)
    mi355x_rows, mi355x_total = parse_mi355x_xlsx(args.mi355x_xlsx)

    print(f"B200: {len(b200_rows)} operators, totals={b200_totals}")
    print(f"MI355X: {len(mi355x_rows)} operators, total_avg={mi355x_total}")

    generate_map(b200_rows, b200_totals, mi355x_rows, mi355x_total,
                 args.b200_csv, args.mi355x_xlsx, args.output)


if __name__ == "__main__":
    main()
