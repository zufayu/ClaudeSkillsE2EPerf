#!/usr/bin/env python3
"""
Extract per-layer kernel breakdown from ATOM trace (ROCm 7.2.2+ compatible).

Replaces parse_trace.py's parse_decode() which requires norm modules.
Uses decode[bs=N] markers and reduce_scatter as layer boundaries.
Output: xlsx matching old parse_trace.py format (2 sheets, module-grouped).

Usage:
    python3 scripts/decode_kernel_breakdown.py <trace.json.gz> \
        --target-bs 64 --layers 10-40
"""

import argparse
import collections
import gzip
import json
import os
import re
import sys

# Import shared trace utilities (R5)
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from trace_utils import load_trace_events  # noqa: E402


# ============================================================================
# Layer structure for DeepSeek-R1 (per transformer layer, execution order):
#
#   input_layernorm:    reduce_scatter + local_device_load_rmsnorm
#   per_token_quant:    dynamic_per_token_scaled_quant  (NEW in rocm722)
#   gemm_a8w8:          gemm_xdl (qkv_a proj)
#   q_k_norm:           fused_qk_rmsnorm_group_quant   (NEW: fused norm+quant)
#   q_proj_and_k_up:    FlatmmKernel/Cijk (q_b) + batched_gemm (k_up)
#   rope_and_kv_cache:  fuse_qk_rope_concat
#   mla_decode:         mla_a8w8 + kn_mla_reduce
#   v_up_proj_o_proj:   batched_gemm (v_up) + gemm_xdl (o_proj)
#   post_attn_layernorm: reduce_scatter + local_device_load_rmsnorm
#   gemm_a16w16:        bf16gemm (router)
#   mxfp4_moe:          grouped_topk + MoeSorting + quant + moe_mxgemm ×2
# ============================================================================

# Positional classifier: maps (kernel_pattern, occurrence_index) → module
# For kernels that appear multiple times per layer, the index disambiguates.
KERNEL_MODULE_RULES = [
    # === GPT-OSS-120B on MI355X (AMD ROCm) ===
    # Must precede generic FlatmmKernel / add_rmsnorm_quant rules below
    # (substring match is first-hit-wins).
    ("MoeFlatmmKernel",                  ["moe_expert_ffn"]),       # gate+up (Swiglu) and down (MoeSilu); down = layer terminator
    ("_fused_add_rmsnorm_pad",           ["input_layernorm"]),
    ("add_rmsnorm_quant_kernel",         ["post_attn_layernorm"]),
    ("_fused_qk_rope_reshape_and_cache", ["rope_and_kv_cache"]),
    ("paged_attention_decode",           ["attention"]),             # covers _sliding_window / _causal variants
    ("topkGatingSoftmax",                ["moe_router"]),
    # (pattern, module_for_1st_occurrence, module_for_2nd_occurrence, ...)
    ("reduce_scatter_cross_device", ["input_layernorm", "post_attn_layernorm"]),
    ("local_device_load_rmsnorm", ["input_layernorm", "post_attn_layernorm"]),
    ("dynamic_per_token_scaled_quant", ["per_token_quant_hip"]),
    ("add_rmsnorm_quant", ["hipLaunchKernel"]),
    ("fused_qk_rmsnorm_group_quant", ["q_proj_and_k_up_proj"]),
    ("gemm_xdl_cshuffle_v3_multi_d_b_preshuffle", ["gemm_a8w8_bpreshuffle", "q_proj_and_k_up_proj"]),
    ("FlatmmKernel", ["v_up_proj_and_o_proj"]),
    ("Cijk_", ["q_proj_and_k_up_proj", "v_up_proj_and_o_proj"]),
    ("batched_gemm_a8w8", ["q_proj_and_k_up_proj", "v_up_proj_and_o_proj"]),
    ("bf16gemm", ["q_proj_and_k_up_proj", "v_up_proj_and_o_proj"]),  # DSR-style 2-occurrence; GPT-OSS has 3rd→clamps to v_up/o_proj
    ("fuse_qk_rope_concat", ["rope_and_kv_cache"]),
    ("mla_a8w8", ["mla_decode"]),
    ("kn_mla_reduce", ["mla_decode"]),
    ("triton_poi_fused", ["triton_poi"]),
    ("grouped_topk_opt_sort", ["mxfp4_moe"]),
    ("MoeSortingMultiPhase", ["mxfp4_moe"]),
    ("mxfp4_quant_moe_sort", ["mxfp4_moe"]),
    ("kernel_moe_mxgemm", ["mxfp4_moe"]),
]


def classify_kernel_positional(kernel_name, occurrence_counters):
    """Classify kernel by name pattern + occurrence count within the layer."""
    for pattern, modules in KERNEL_MODULE_RULES:
        if pattern in kernel_name:
            idx = occurrence_counters.get(pattern, 0)
            occurrence_counters[pattern] = idx + 1
            return modules[min(idx, len(modules) - 1)]
    return "other"


def load_trace(path):
    """Delegates to trace_utils.load_trace_events."""
    return load_trace_events(path)


def select_decodes(evts, target_bs, skip_warmup=5, max_steps=20):
    """Return a list of decode events to aggregate over.

    Skips the first `skip_warmup` decodes (cold caches / JIT) and caps at
    `max_steps`. Replaces the old `select_decode` (single-decode picker) so
    we can aggregate ~600 (step,layer) samples per operator instead of ~30.
    Convention matches B300/B200 trace_layer_detail.py for cross-platform
    apples-to-apples comparison.
    """
    # cat='gpu_user_annotation' filter is critical: without it we also catch
    # capture-graph replay events that look like decode[bs=N] but are sub-ms
    # (graph capture / MTP draft). With it, we only get the real decode
    # iterations (typically 10-30ms each on MI355X).
    decodes = sorted(
        [
            e for e in evts
            if e.get("name", "").startswith("decode[")
            and e.get("ph") == "X"
            and e.get("cat") == "gpu_user_annotation"
            and f"bs={target_bs}" in e.get("name", "")
        ],
        key=lambda x: x["ts"],
    )
    if not decodes:
        print(f"ERROR: no decode events with bs={target_bs} (cat=gpu_user_annotation)")
        return []
    if len(decodes) <= skip_warmup:
        print(f"WARN: only {len(decodes)} decodes — can't skip {skip_warmup} warmup; using all")
        return decodes
    selected = decodes[skip_warmup : skip_warmup + max_steps] if max_steps > 0 \
        else decodes[skip_warmup:]
    print(f"  decode events (bs={target_bs}): {len(decodes)} total")
    print(f"  Selected {len(selected)} steady-state decodes "
          f"(skipped first {skip_warmup} warmup, cap {max_steps})")
    return selected


def extract_kernels_in_window(evts, ts_start, ts_end):
    return sorted(
        [e for e in evts if e.get("cat") == "kernel" and e.get("ph") == "X"
         and e["ts"] >= ts_start and e["ts"] <= ts_end],
        key=lambda e: e["ts"]
    )


# ----------------------------------------------------------------------------
# Robust per-layer anchor detection (R8d — Task #11, structural / name-agnostic)
# ----------------------------------------------------------------------------
# Attention kernel names differ across models/backends and sometimes vary
# *within* a model (dynamic shapes, FlashInfer autotune, ATOM jit-fused names).
# Hardcoding name patterns breaks. Instead use a structural property:
# every transformer layer fires the SAME named GPU kernel exactly once per
# decode iteration, so the layer-anchor is the kernel that:
#   (a) fires `layer_count` times in one decode window
#   (b) fires at regular intervals (low CV of inter-fire gaps)
# `layer_count` itself is auto-discovered (= the count of the chosen anchor).
#
# Reference layer counts for common open models (sanity check the printed
# anchor count against these — wildly off means the heuristic mispicked):
#   DeepSeek-R1 / V3 / V2.5  61    Llama 3.x 70B          80
#   DeepSeek-V2              60    Llama 3.1 405B        126
#   Llama 2/3 7B/8B          32    Llama 2 13B / Phi-3 14B  40
#   Mistral 7B / Mixtral 8x7B 32   Mixtral 8x22B           56
#   Qwen 2.5 7B              28    Qwen 2.5 14B            48
#   Qwen 2.5 32B             64    Qwen 2.5/3 72B/32B      80/64
#   Qwen 3 235B              94    Gemma 2 9B / 27B       42 / 46
# If the printed count looks wrong, override with --attn-kernel <regex> or
# tighten --anchor-min/--anchor-max bounds.


def _stdev(xs):
    if len(xs) < 2:
        return 0.0
    m = sum(xs) / len(xs)
    return (sum((x - m) ** 2 for x in xs) / len(xs)) ** 0.5


def find_layer_anchor(kernels, anchor_min=20, anchor_max=150,
                      max_cv=0.5, override_regex=None):
    """Detect the per-layer anchor kernel structurally (no name assumptions).

    Returns (anchor_name, [positions]). Raises ValueError if nothing fits.

    Algorithm:
      1. Group kernel positions by exact name.
      2. Discard names whose count is outside [anchor_min, anchor_max] (typical
         model layer count is 28-126; 20-150 range covers everything).
      3. For each remaining name, compute CV of inter-fire intervals.
      4. Discard names with CV > max_cv (irregular spacing → not a per-layer
         anchor; e.g. element-wise ops fire on demand, not periodically).
      5. Pick the name with lowest interval CV (most regular). Tiebreak: count
         closest to median of plausible model layer counts (~50).

    --attn-kernel REGEX bypasses everything and uses regex matches as anchor.
    """
    if override_regex:
        positions = [i for i, k in enumerate(kernels)
                     if re.search(override_regex, k.get("name", ""), re.IGNORECASE)]
        if not positions:
            raise ValueError(
                f"--attn-kernel regex {override_regex!r} matched 0 kernels in decode")
        return override_regex, positions

    # Group by exact name (note: dynamic-shape kernels with templated names
    # are still grouped — they share the same emitted name string)
    name_positions = {}
    for i, k in enumerate(kernels):
        name = k.get("name", "")
        if not name:
            continue
        name_positions.setdefault(name, []).append(i)

    # Filter to plausible layer-count range
    candidates = []
    for name, positions in name_positions.items():
        n = len(positions)
        if not (anchor_min <= n <= anchor_max):
            continue
        intervals = [positions[i+1] - positions[i] for i in range(len(positions)-1)]
        if not intervals:
            continue
        mean_int = sum(intervals) / len(intervals)
        if mean_int <= 0:
            continue
        cv = _stdev(intervals) / mean_int
        if cv > max_cv:
            continue
        candidates.append((cv, abs(n - 50), name, positions))

    if not candidates:
        # Build a diagnostic showing top-5 most-frequent kernels in plausible range
        debug = sorted(
            ((len(p), name) for name, p in name_positions.items()
             if anchor_min <= len(p) <= anchor_max),
            reverse=True,
        )[:5]
        raise ValueError(
            f"No structural per-layer anchor found (count in [{anchor_min},{anchor_max}], "
            f"interval CV ≤ {max_cv}). Top frequency candidates (count, name): {debug}. "
            f"Override with --attn-kernel <regex> if you know the right kernel.")

    # Lowest CV first, then count closest to ~50 (heuristic median model layer count)
    candidates.sort()
    cv, _, best_name, best_pos = candidates[0]
    return best_name, best_pos


def split_layers(kernels, anchor_positions=None, override_regex=None):
    """Split kernels into per-layer windows using anchor-to-anchor cuts.

    Each layer = kernels[anchor[i] : anchor[i+1]]. This is name-agnostic and
    works for any model (no reliance on reduce_scatter / specific kernel names
    that vary across ATOM/sglang/vLLM/TRT-LLM and across model architectures).

    Note: position 0 within each layer is the anchor kernel itself (e.g. attn).
    Pre-attn kernels (norm/quant/qkv_a) end up at the END of the PREVIOUS
    layer's window. The phase shift is constant across all decodes, so per-
    position aggregation is consistent (just remember the indexing convention).

    Args:
        kernels: GPU kernels in one decode window, sorted by ts.
        anchor_positions: pre-computed list (from find_layer_anchor). If None,
                          auto-detect now.
        override_regex: explicit anchor pattern when auto-detecting.

    Returns: list of layer-kernel-lists (length = len(anchor_positions)).
    """
    if anchor_positions is None:
        try:
            _, anchor_positions = find_layer_anchor(kernels, override_regex=override_regex)
        except ValueError:
            return [kernels]

    if len(anchor_positions) < 2:
        return [kernels]

    layers = []
    for i in range(len(anchor_positions) - 1):
        layers.append(kernels[anchor_positions[i]:anchor_positions[i+1]])
    # Tail after last anchor — usually a partial layer (just attn through end of decode);
    # keep it so callers can see it but they typically drop incomplete layers.
    layers.append(kernels[anchor_positions[-1]:])
    return layers


def classify_layer(layer_kernels):
    """Classify each kernel in a layer using positional rules. Returns [(module, kernel_name, dur_us)]."""
    counters = {}
    result = []
    for k in layer_kernels:
        name = k.get("name", "unknown")
        module = classify_kernel_positional(name, counters)
        result.append((module, name, k.get("dur", 0)))
    return result


def parse_layer_spec(spec):
    """Parse --layers arg into a sorted list of int indices.

    Forms:
      'lo-hi'           -> [lo, lo+1, ..., hi]   (inclusive both ends)
      'lo-hi:even'      -> even indices in lo..hi
      'lo-hi:odd'       -> odd indices in lo..hi
      'i,j,k,...'       -> exactly those indices
    """
    spec = spec.strip()
    if "," in spec:
        idxs = [int(x) for x in spec.split(",") if x.strip()]
    else:
        parity = None
        if ":" in spec:
            spec, parity = spec.split(":", 1)
            parity = parity.strip().lower()
            if parity not in ("even", "odd"):
                raise ValueError(f"--layers parity must be 'even' or 'odd', got {parity!r}")
        lo, hi = map(int, spec.split("-"))
        idxs = list(range(lo, hi + 1))
        if parity == "even":
            idxs = [i for i in idxs if i % 2 == 0]
        elif parity == "odd":
            idxs = [i for i in idxs if i % 2 == 1]
    if not idxs:
        raise ValueError(f"--layers {spec!r} resolved to empty set")
    return sorted(set(idxs))


def main():
    parser = argparse.ArgumentParser(description="Decode kernel breakdown from ATOM trace")
    parser.add_argument("trace", help="Path to trace JSON or JSON.GZ")
    parser.add_argument("--target-bs", type=int, default=64)
    parser.add_argument("--skip-warmup", type=int, default=5,
                        help="Drop first N decode steps (cold caches / JIT). "
                             "Matches B300 trace_layer_detail convention.")
    parser.add_argument("--max-steps", type=int, default=20,
                        help="Aggregate over N steady-state decodes (default 20 → "
                             "~20 × 30 layers = 600 samples per operator). 0 = all available.")
    parser.add_argument("--skip-ratio", type=float, default=None,
                        help="DEPRECATED. Old single-decode picker. Ignored when "
                             "--skip-warmup / --max-steps are used.")
    parser.add_argument("--layers", type=str, default="10-40",
                        help="Layer index selection. Forms: 'lo-hi' (inclusive range), "
                             "'lo-hi:even' or 'lo-hi:odd' (parity filter — useful for "
                             "GPT-OSS-style models that alternate sliding-window vs "
                             "full-context attention 1:1 by layer parity), or a comma "
                             "list 'i,j,k,...'. Default 10-40 = 30 layers (steady "
                             "middle, common for 60+ layer models like DSR1/Llama70B). "
                             "For smaller models (Llama 7B 32 layers, Mistral 7B 32 "
                             "layers), use e.g. 5-25.")
    parser.add_argument("--anchor-min", type=int, default=20,
                        help="Min plausible per-decode layer count (for structural "
                             "anchor detection). Default 20 covers smallest open models.")
    parser.add_argument("--anchor-max", type=int, default=150,
                        help="Max plausible per-decode layer count. Default 150 covers "
                             "Llama-405B (126).")
    parser.add_argument("--attn-kernel", type=str, default=None,
                        help="Explicit anchor-kernel regex (skip auto-detect). Use when "
                             "auto-detection picks a wrong kernel — e.g. force "
                             "'mla_a8w8' on ATOM int8 MLA traces.")
    parser.add_argument("--enforce-min-samples", type=int, default=0,
                        help="Fail with non-zero exit if final aggregated layer count "
                             "drops below this (e.g. 500 to require ~17 of 20 decodes "
                             "× 30 layers). Use in CI to catch silent under-sampling. "
                             "Default 0 = no enforcement.")
    parser.add_argument("--output", type=str, default=None)
    args = parser.parse_args()

    if args.output is None:
        m = re.search(r'_(c\d+)[_.]', os.path.basename(args.trace))
        suffix = m.group(1) if m else f"c{args.target_bs}"
        args.output = f"decode_breakdown_{suffix}.xlsx"

    layer_indices = parse_layer_spec(args.layers)
    layer_start = layer_indices[0]
    layer_end = layer_indices[-1]
    layer_set = set(layer_indices)
    evts = load_trace(args.trace)

    decodes = select_decodes(evts, args.target_bs, args.skip_warmup, args.max_steps)
    if not decodes:
        sys.exit(1)

    # Pool classified layers across all selected decodes (R8d wide-sample).
    # Per-decode quality gate: reject decodes whose anchor count is too low to
    # cover layer_start..layer_end (otherwise per-decode contribution would be
    # truncated, polluting per-position aggregation).
    all_classified = []        # list of [(module, kname, dur), ...] per layer
    total_dur_ms = 0.0
    n_used = 0
    n_rejected = 0
    rejection_log = []
    detected_anchor = None     # cache the anchor name from first successful decode
    for i, decode in enumerate(decodes):
        ts0, dur = decode["ts"], decode.get("dur", 0)
        kernels = extract_kernels_in_window(evts, ts0, ts0 + dur)

        # Find layer anchor for this decode (structural detection, name-agnostic)
        try:
            anchor_name, anchor_pos = find_layer_anchor(
                kernels,
                anchor_min=args.anchor_min,
                anchor_max=args.anchor_max,
                override_regex=args.attn_kernel,
            )
        except ValueError as e:
            n_rejected += 1
            rejection_log.append(f"  decode #{i:02d} ts={ts0}: anchor fail — {str(e)[:200]}")
            continue

        if i == 0 or detected_anchor is None:
            detected_anchor = anchor_name
            print(f"  Anchor: {anchor_name!r} ({len(anchor_pos)} positions in decode #{i})")

        # Reject if anchor count < required (need at least layer_end+1 anchors
        # to slice up to layer index layer_end)
        min_required = layer_end + 1
        if len(anchor_pos) < min_required:
            n_rejected += 1
            rejection_log.append(
                f"  decode #{i:02d}: only {len(anchor_pos)} anchors (need ≥{min_required} "
                f"for layers {layer_start}-{layer_end})")
            continue

        layers = split_layers(kernels, anchor_positions=anchor_pos)
        # Skip partial first layer (may start mid-layer in post-attn phase)
        if len(layers) > 1:
            first_names = [k.get("name", "") for k in layers[0][:5]]
            if any(any(p in n for p in ["bf16gemm", "grouped_topk", "kernel_moe"])
                   for n in first_names):
                layers = layers[1:]
        # Pick exactly the requested indices (supports range, parity, list)
        n_layers = len(layers)
        selected = [layers[i] for i in layer_indices if 0 <= i < n_layers]
        for l in selected:
            all_classified.append(classify_layer(l))
        total_dur_ms += dur / 1000
        n_used += 1

    n_layers_total = len(all_classified)
    n_decodes = len(decodes)
    avg_dur_ms = total_dur_ms / n_used if n_used else 0
    print(f"\n  Aggregated {n_layers_total} layer samples across {n_used}/{n_decodes} decodes "
          f"({n_rejected} rejected)")
    if rejection_log:
        print(f"  Rejection log (first 5):")
        for r in rejection_log[:5]:
            print(r)
    print(f"  Avg decode walltime: {avg_dur_ms:.2f}ms")

    if not all_classified:
        print(f"ERROR: no layers extracted from {n_decodes} decodes — check --target-bs, "
              f"--layers, --attn-kernel. Anchor detected: {detected_anchor!r}.")
        sys.exit(1)

    # R8d guard rail: --enforce-min-samples
    if args.enforce_min_samples > 0 and n_layers_total < args.enforce_min_samples:
        print(f"ERROR: --enforce-min-samples={args.enforce_min_samples} not met "
              f"(got {n_layers_total} layer samples). "
              f"Causes: short trace, too-strict warmup, or anchor-count mismatch.")
        sys.exit(2)

    # Representative layer (for xlsx Sheet 1 single-layer column display only —
    # the per-position aggregate stats below are what actually matter).
    rep_layer = all_classified[len(all_classified) // 2]

    # Per-position aggregates (mean, median, p95) across all (decode × layer) samples
    max_pos = max(len(cl) for cl in all_classified)
    avg_durs = []
    med_durs = []
    p95_durs = []
    for pos in range(max_pos):
        durs = sorted(cl[pos][2] for cl in all_classified if pos < len(cl))
        if not durs:
            avg_durs.append(0); med_durs.append(0); p95_durs.append(0)
            continue
        avg_durs.append(sum(durs) / len(durs))
        med_durs.append(durs[len(durs) // 2])
        p95_idx = max(0, int(len(durs) * 0.95) - 1)
        p95_durs.append(durs[p95_idx])

    total_rep = sum(d for _, _, d in rep_layer)
    total_avg = sum(avg_durs[:len(rep_layer)])
    n_layers = n_layers_total  # for back-compat below

    print(f"\n{'='*60}")
    print(f"LAYER {args.layers} ANALYSIS "
          f"({n_layers_total} layer samples × {n_decodes} decodes)")
    print(f"{'='*60}")
    print(f"  Representative layer: {len(rep_layer)} kernels")
    print(f"  Representative-layer total: {total_rep:.1f}us ({total_rep/1000:.2f}ms)")
    print(f"  Pooled-mean per-layer total: {total_avg:.1f}us ({total_avg/1000:.2f}ms)")
    print(f"  Est decode (x61): {total_avg * 61 / 1000:.1f}ms (actual avg: {avg_dur_ms:.2f}ms)")

    # Print layer structure
    print(f"\n{'#':<3} {'Module':<40} {'Kernel':<55} {'dur(us)':>8} {'avg':>8}")
    print("-" * 120)
    cur_mod = ""
    for i, (mod, kname, d) in enumerate(rep_layer):
        avg = avg_durs[i] if i < len(avg_durs) else 0
        show_mod = mod if mod != cur_mod else ""
        if mod != cur_mod:
            cur_mod = mod
        print(f"  {i:<3} {show_mod:<40} {kname[:53]:<55} {d:>7.1f} {avg:>7.1f}")

    # Partition representative-layer kernels into attention vs non-attention.
    # Attention block in a transformer layer is contiguous (norm -> q/k/v ->
    # rope+kv -> attn -> o_proj), so this split preserves intra-block order.
    def is_attn_kernel(name):
        nl = name.lower()
        return any(s in nl for s in (
            "paged_attention", "fmha", "mla_", "_fused_qk_rope",
            "flash_attn", "attention_kernel",
        ))

    # Write xlsx with 4 sheets: attn_decode, attn_summary, non_attn_decode,
    # non_attn_summary. Each {prefix}_decode is sequential / module-grouped;
    # each {prefix}_summary is aggregated by kernel name.
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    def emit_sheets(title_prefix, indices):
        sub = [(rep_layer[i], avg_durs[i], med_durs[i], p95_durs[i]) for i in indices]
        sub_total_rep = sum(d for (_, _, d), *_ in sub)
        sub_total_avg = sum(a for _, a, _, _ in sub)

        ws = wb.create_sheet(f"{title_prefix}_decode")
        ws.append(["cpu_module", "gpu_kernel", "duration_us", "pct%",
                   "sum per module", "module_pct%",
                   "avg_us", "median_us", "p95_us", "avg sum per module",
                   "n_samples"])
        groups = []
        for (mod, kname, d), avg, med, p95 in sub:
            if not groups or groups[-1][0] != mod:
                groups.append((mod, []))
            groups[-1][1].append((kname, d, avg, med, p95))
        for mod, kernel_list in groups:
            mod_sum = sum(d for _, d, _, _, _ in kernel_list)
            mod_avg_sum = sum(a for _, _, a, _, _ in kernel_list)
            mod_pct = mod_sum / sub_total_rep * 100 if sub_total_rep else 0
            first = True
            for kname, d, avg, med, p95 in kernel_list:
                pct = d / sub_total_rep * 100 if sub_total_rep else 0
                ws.append([
                    mod if first else "",
                    kname,
                    round(d, 3),
                    round(pct, 1),
                    round(mod_sum, 3) if first else "",
                    round(mod_pct, 1) if first else "",
                    round(avg, 3),
                    round(med, 3),
                    round(p95, 3),
                    round(mod_avg_sum, 3) if first else "",
                    n_layers_total,
                ])
                first = False
        ws.append(["TOTAL", "", round(sub_total_rep, 3), "100", "", "100",
                   round(sub_total_avg, 3), "", "", "", n_layers_total])

        ws2 = wb.create_sheet(f"{title_prefix}_summary")
        ws2.append(["gpu_kernel", "calls", "total_duration_us",
                    "avg_duration_us", "pct%"])
        by_name = collections.defaultdict(lambda: {"count": 0, "dur": 0})
        for (_mod, kname, d), *_ in sub:
            by_name[kname]["count"] += 1
            by_name[kname]["dur"] += d
        for kname, v in sorted(by_name.items(), key=lambda x: -x[1]["dur"]):
            pct = v["dur"] / sub_total_rep * 100 if sub_total_rep else 0
            avg_d = v["dur"] / v["count"] if v["count"] else 0
            ws2.append([kname, v["count"], round(v["dur"], 1),
                        round(avg_d, 1), round(pct, 1)])
        return len(groups), sub_total_rep

    attn_idx = [i for i, (_, kn, _) in enumerate(rep_layer) if is_attn_kernel(kn)]
    non_attn_idx = [i for i in range(len(rep_layer)) if i not in set(attn_idx)]
    n_attn_groups, attn_total = emit_sheets("attn", attn_idx)
    n_non_groups, non_attn_total = emit_sheets("non_attn", non_attn_idx)

    print(f"\nWriting {args.output}...")
    wb.save(args.output)
    print(f"Done. attn: {len(attn_idx)} kernels / {n_attn_groups} groups "
          f"({attn_total:.1f}us). "
          f"non_attn: {len(non_attn_idx)} kernels / {n_non_groups} groups "
          f"({non_attn_total:.1f}us).")


if __name__ == "__main__":
    main()
